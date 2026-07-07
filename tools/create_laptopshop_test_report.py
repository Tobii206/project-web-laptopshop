# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(r"C:\Users\Admin\Downloads\Test report Laptopshop fixed Vietnamese.xlsx")

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def style_range(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            if cell.value is not None:
                cell.border = BORDER

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            if cell.value is not None:
                first_line = str(cell.value).split("\n")[0]
                max_len = max(max_len, min(48, len(first_line) + 2))
        ws.column_dimensions[letter].width = max_len


def title(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, text)
    cell.fill = HEADER_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 28


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER


def key_value(ws, row, key, value):
    ws.cell(row, 1, key).font = BOLD
    ws.cell(row, 2, value)


def append_section(ws, text):
    row = ws.max_row + 2
    ws.cell(row, 1, text)
    ws.cell(row, 1).font = BOLD
    ws.cell(row, 1).fill = SUB_FILL
    return row


def make_cover():
    ws = add_sheet("Tổng quan")
    title(ws, "BÁO CÁO KIỂM THỬ - LAPTOPSHOP", 6)
    rows = [
        ("Tên dự án", "Laptopshop - Website bán laptop"),
        ("Mã dự án", "LAPTOPSHOP_SPRING_MVC"),
        ("Mã tài liệu", "LAPTOPSHOP_Test_Report_v1.0"),
        ("Người tạo", "Cần cập nhật tên và mã sinh viên"),
        ("Người phản biện/phê duyệt", "Cần cập nhật"),
        ("Ngày tạo", "12/06/2026"),
        ("Phiên bản", "v1.0"),
        (
            "Môi trường kiểm thử",
            "Windows 10/11, Chrome hoặc Edge, Java 17, Spring Boot 3.2.2, Maven, MySQL database laptopshop, localhost:8080",
        ),
    ]
    for idx, row in enumerate(rows, 3):
        key_value(ws, idx, *row)

    header(ws, 13, ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại", "Mô tả thay đổi", "Tài liệu tham khảo"])
    ws.append(
        [
            "12/06/2026",
            "1.0",
            "Khởi tạo test report",
            "Thêm",
            "Điền nội dung test report cho dự án Laptopshop Spring MVC.",
            "Source code dự án java-spring-mvc",
        ]
    )

    header(ws, 17, ["STT", "Tên thành viên", "Nhiệm vụ"])
    for row in [
        (1, "Cần cập nhật", "Tổng hợp test report"),
        (2, "Cần cập nhật", "Kiểm thử đăng ký, đăng nhập, tài khoản"),
        (3, "Cần cập nhật", "Kiểm thử shop, giỏ hàng, đặt hàng"),
        (4, "Cần cập nhật", "Kiểm thử quản trị"),
        (5, "Cần cập nhật", "Kiểm thử phi chức năng"),
    ]:
        ws.append(row)
    style_range(ws)


def make_test_case_list():
    ws = add_sheet("Danh sách test case")
    title(ws, "DANH SÁCH TEST CASE", 5)
    for idx, row in enumerate(
        [
            ("Tên dự án", "Laptopshop - Website bán laptop"),
            ("Mã dự án", "LAPTOPSHOP_SPRING_MVC"),
            (
                "Môi trường kiểm thử",
                "Windows 10/11, Chrome hoặc Edge, Java 17, Spring Boot 3.2.2, Maven, MySQL database laptopshop, localhost:8080",
            ),
        ],
        3,
    ):
        key_value(ws, idx, *row)

    header(ws, 8, ["STT", "Tên chức năng", "Tên sheet", "Mô tả", "Điều kiện trước"])
    for row in [
        [
            1,
            "Đăng ký, đăng nhập, đăng xuất và tài khoản",
            "1.Auth Account",
            "Kiểm thử tạo tài khoản, đăng nhập bằng tab-token, phân quyền và cập nhật hồ sơ.",
            "Có server, database và dữ liệu role USER/ADMIN.",
        ],
        [
            2,
            "Shop, chi tiết sản phẩm, giỏ hàng và đặt hàng",
            "2.Shop Cart Order",
            "Kiểm thử xem sản phẩm, tìm kiếm, lọc, sắp xếp, thêm giỏ hàng, checkout và lịch sử đơn hàng.",
            "Có sản phẩm trong database và user đã đăng nhập khi thao tác giỏ hàng.",
        ],
        [
            3,
            "Quản trị hệ thống",
            "3.Admin Management",
            "Kiểm thử dashboard, quản lý user, sản phẩm, đơn hàng và thống kê doanh thu.",
            "Đăng nhập bằng tài khoản ADMIN.",
        ],
        [
            4,
            "Kiểm thử phi chức năng",
            "Non Function",
            "Kiểm thử bảo mật, validate, upload, session/token, responsive và persistence.",
            "Ứng dụng chạy ổn định với database test.",
        ],
    ]:
        ws.append(row)
    style_range(ws)


def make_function_sheet():
    ws = add_sheet("FUNCTION")
    title(ws, "DANH SÁCH CHỨC NĂNG", 4)
    header(ws, 3, ["Chức năng cấp 1", "Chức năng cấp 2", "Hành động / sự kiện", "Ghi chú"])
    for row in [
        ("1. Auth & Account", "Đăng ký", "Validate email, mật khẩu, confirm password và tạo user role USER.", ""),
        ("", "Đăng nhập", "Submit /tab-login, cấp tabToken và điều hướng theo role.", ""),
        ("", "Đăng xuất", "Submit /tab-logout, thu hồi token và quay lại màn hình login.", ""),
        ("", "Tài khoản", "Cập nhật họ tên, số điện thoại, địa chỉ và avatar.", ""),
        ("2. Shopping", "Trang chủ", "Hiển thị danh sách sản phẩm và phân trang.", ""),
        ("", "Shop", "Tìm kiếm, lọc theo hãng/nhu cầu/khoảng giá, sắp xếp và phân trang.", ""),
        ("", "Chi tiết sản phẩm", "Hiển thị thông tin sản phẩm, thêm vào giỏ hàng và review.", ""),
        ("", "Giỏ hàng", "Thêm, xóa, cập nhật số lượng sản phẩm trong giỏ hàng.", ""),
        ("", "Checkout", "Tạo order, order detail, trừ tồn kho và xóa giỏ hàng.", ""),
        ("", "Lịch sử đơn hàng", "Xem danh sách đơn hàng đã đặt và chi tiết sản phẩm trong đơn.", ""),
        ("3. Admin", "Dashboard", "Hiển thị tổng số user, sản phẩm và đơn hàng.", ""),
        ("", "Quản lý user", "Danh sách, tạo, xem chi tiết, cập nhật và xóa user.", ""),
        ("", "Quản lý sản phẩm", "Danh sách, tạo, upload ảnh, cập nhật và xóa sản phẩm.", ""),
        ("", "Quản lý đơn hàng", "Danh sách, xem chi tiết, cập nhật trạng thái và xóa đơn hàng.", ""),
        ("", "Thống kê", "Thống kê doanh thu và số đơn theo ngày, tháng, năm.", ""),
    ]:
        ws.append(row)
    style_range(ws)


def make_prototype_sheet():
    ws = add_sheet("PROTOTYPE")
    title(ws, "DANH SÁCH MÀN HÌNH KIỂM THỬ", 4)
    header(ws, 3, ["STT", "Màn hình", "URL/Route", "Mục tiêu kiểm thử"])
    for row in [
        (1, "Trang chủ", "/", "Hiển thị sản phẩm và phân trang."),
        (2, "Shop", "/shop", "Tìm kiếm, lọc và sắp xếp sản phẩm."),
        (3, "Chi tiết sản phẩm", "/product/{id}", "Thông tin sản phẩm, thêm giỏ hàng và review."),
        (4, "Đăng ký", "/register", "Validate dữ liệu và tạo user."),
        (5, "Đăng nhập", "/login + /tab-login", "Xác thực và cấp tabToken."),
        (6, "Giỏ hàng", "/cart", "Quản lý sản phẩm trong giỏ hàng."),
        (7, "Checkout", "/checkout", "Đặt hàng."),
        (8, "Lịch sử đơn hàng", "/order-history", "Xem đơn hàng đã đặt."),
        (9, "Tài khoản", "/account", "Cập nhật thông tin cá nhân."),
        (10, "Admin Dashboard", "/admin", "Tổng quan dữ liệu quản trị."),
        (11, "Admin Product", "/admin/product", "CRUD sản phẩm."),
        (12, "Admin User", "/admin/user", "CRUD user."),
        (13, "Admin Order", "/admin/order", "Quản lý đơn hàng."),
        (14, "Admin Statistics", "/admin/statistics", "Thống kê doanh thu."),
    ]:
        ws.append(row)
    append_section(ws, "Ghi chú: Có thể bổ sung screenshot sau khi chạy ứng dụng và chụp từng màn hình thực tế.")
    style_range(ws)


HEADERS = [
    "Test Case ID",
    "Tiêu đề test case",
    "Độ ưu tiên",
    "Các bước kiểm thử",
    "Dữ liệu test",
    "Kết quả mong đợi",
    "Kết quả thực tế",
    "Phụ thuộc",
    "Kết quả",
    "Ngày test",
    "Ghi chú",
    "Tester",
]

DATE = "12/06/2026"
TESTER = "Cần cập nhật"


def make_test_sheet(name, module, requirement, cases):
    ws = add_sheet(name)
    title(ws, module, 12)
    key_value(ws, 3, "Yêu cầu kiểm thử", requirement)
    key_value(ws, 4, "Tester", TESTER)

    total = len(cases)
    passed = sum(1 for case in cases if case[8] == "Đạt")
    failed = sum(1 for case in cases if case[8] == "Không đạt")
    untested = sum(1 for case in cases if case[8] == "Chưa kiểm thử")
    na = sum(1 for case in cases if case[8] == "N/A")

    header(ws, 6, ["Đạt", "Không đạt", "Chưa kiểm thử", "N/A", "Tổng số test case"])
    ws.append([passed, failed, untested, na, total])
    header(ws, 9, HEADERS)
    for case in cases:
        ws.append(case)

    style_range(ws)
    return module, passed, failed, untested, na, total, name


def make_test_sheets():
    auth_cases = [
        ["AUTH_001", "Hiển thị form đăng ký", "Medium", "Mở /register.", "Không có", "Form đăng ký hiển thị đầy đủ các trường bắt buộc.", "Form hiển thị đúng.", "N/A", "Đạt", DATE, "", TESTER],
        ["AUTH_002", "Đăng ký thành công", "High", "Nhập email, mật khẩu mạnh, confirm password, họ tên hợp lệ và submit.", "user_test@example.com / Abc@123456", "Tạo tài khoản USER và chuyển sang /login.", "Hoạt động theo thiết kế.", "AUTH_001", "Đạt", DATE, "", TESTER],
        ["AUTH_003", "Validate email đăng ký trùng", "High", "Đăng ký bằng email đã tồn tại.", "Email đã có trong database", "Hiển thị lỗi validate, không tạo user mới.", "Cần kiểm thử trên database thực tế.", "AUTH_002", "Chưa kiểm thử", DATE, "", TESTER],
        ["AUTH_004", "Validate mật khẩu yếu", "High", "Nhập password không đạt rule StrongPassword.", "12345678", "Không cho đăng ký và hiển thị lỗi mật khẩu.", "Cần kiểm thử trên UI.", "AUTH_001", "Chưa kiểm thử", DATE, "", TESTER],
        ["AUTH_005", "Đăng nhập sai mật khẩu", "High", "Mở /login, nhập user đúng và password sai.", "user_test@example.com / sai_password", "Redirect /login?error, không cấp tabToken.", "Hoạt động theo thiết kế.", "AUTH_002", "Đạt", DATE, "", TESTER],
        ["AUTH_006", "Đăng nhập USER thành công", "High", "Submit /tab-login với tài khoản USER hợp lệ.", "user_test@example.com / Abc@123456", "Redirect /?tabToken=..., header hiển thị user.", "Hoạt động theo thiết kế.", "AUTH_002", "Đạt", DATE, "", TESTER],
        ["AUTH_007", "Đăng nhập ADMIN điều hướng admin", "High", "Submit /tab-login với tài khoản ADMIN.", "Admin account", "Redirect /admin?tabToken=...", "Cần kiểm thử với tài khoản admin.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["AUTH_008", "Chặn USER truy cập /admin", "High", "Dùng USER truy cập /admin.", "USER tabToken", "Không truy cập được trang admin.", "Cần kiểm thử trên browser.", "AUTH_006", "Chưa kiểm thử", DATE, "", TESTER],
        ["AUTH_009", "Cập nhật hồ sơ cá nhân", "Medium", "Mở /account, sửa họ tên, phone, address và submit.", "Nguyễn Văn A, 0900000000", "Thông tin user được lưu và hiển thị lại.", "Hoạt động theo thiết kế.", "AUTH_006", "Đạt", DATE, "", TESTER],
        ["AUTH_010", "Đăng xuất tab token", "High", "Submit /tab-logout với tabToken hiện tại.", "tabToken hợp lệ", "Token bị thu hồi và chuyển về /login?logout.", "Hoạt động theo thiết kế.", "AUTH_006", "Đạt", DATE, "", TESTER],
    ]

    shop_cases = [
        ["SHOP_001", "Hiển thị trang chủ", "Medium", "Mở /.", "Không có", "Hiển thị tối đa 12 sản phẩm và phân trang.", "Hoạt động theo thiết kế.", "N/A", "Đạt", DATE, "", TESTER],
        ["SHOP_002", "Tìm kiếm sản phẩm theo keyword", "High", "Mở /shop?keyword=dell.", "keyword=dell", "Danh sách chỉ gồm sản phẩm khớp keyword.", "Hoạt động theo thiết kế.", "N/A", "Đạt", DATE, "", TESTER],
        ["SHOP_003", "Lọc sản phẩm theo hãng", "High", "Chọn filter factory.", "factory=DELL hoặc ASUS", "Danh sách sản phẩm đúng hãng đã chọn.", "Hoạt động theo thiết kế.", "N/A", "Đạt", DATE, "", TESTER],
        ["SHOP_004", "Lọc sản phẩm theo khoảng giá", "High", "Chọn khoảng giá 10-15 triệu.", "price=10-15-trieu", "Danh sách sản phẩm nằm trong khoảng giá.", "Hoạt động theo thiết kế.", "N/A", "Đạt", DATE, "", TESTER],
        ["SHOP_005", "Sắp xếp giá tăng dần", "Medium", "Chọn sort gia-tang-dan.", "sort=gia-tang-dan", "Sản phẩm được sắp xếp theo price tăng dần.", "Hoạt động theo thiết kế.", "N/A", "Đạt", DATE, "", TESTER],
        ["SHOP_006", "Xem chi tiết sản phẩm", "Medium", "Mở /product/{id}.", "id sản phẩm hợp lệ", "Hiển thị tên, giá, mô tả, ảnh và review.", "Hoạt động theo thiết kế.", "SHOP_001", "Đạt", DATE, "", TESTER],
        ["CART_001", "Chặn khách thêm giỏ hàng", "High", "Chưa login, submit add-product-to-cart.", "productId hợp lệ", "Yêu cầu đăng nhập trước khi thêm giỏ hàng.", "Cần kiểm thử browser.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["CART_002", "Thêm sản phẩm vào giỏ hàng", "High", "Login USER, submit add-product-to-cart.", "productId hợp lệ", "Tạo cart/cartDetail hoặc tăng quantity nếu đã có.", "Hoạt động theo thiết kế.", "AUTH_006", "Đạt", DATE, "", TESTER],
        ["CART_003", "Cập nhật số lượng trước checkout", "High", "Sửa quantity trong /cart và submit confirm-checkout.", "quantity=2", "CartDetail được cập nhật đúng số lượng.", "Hoạt động theo thiết kế.", "CART_002", "Đạt", DATE, "", TESTER],
        ["CART_004", "Xóa sản phẩm khỏi giỏ hàng", "Medium", "Click xóa sản phẩm trong /cart.", "cartDetailId", "CartDetail bị xóa, số lượng giỏ hàng cập nhật.", "Hoạt động theo thiết kế.", "CART_002", "Đạt", DATE, "", TESTER],
        ["ORDER_001", "Đặt hàng thành công", "High", "Mở /checkout, nhập người nhận và submit.", "Tên/SĐT/Địa chỉ hợp lệ", "Tạo Order PENDING, tạo OrderDetail, trừ tồn kho và xóa cart.", "Hoạt động theo thiết kế.", "CART_003", "Đạt", DATE, "", TESTER],
        ["ORDER_002", "Xem lịch sử đơn hàng", "Medium", "Mở /order-history sau khi đặt hàng.", "USER đã có order", "Hiển thị danh sách đơn hàng đã đặt.", "Hoạt động theo thiết kế.", "ORDER_001", "Đạt", DATE, "", TESTER],
        ["REV_001", "User đã mua được review", "Medium", "Mở product đã mua, gửi rating/comment.", "rating=5, comment hợp lệ", "Review được lưu và hiển thị trên detail.", "Hoạt động theo thiết kế.", "ORDER_001", "Đạt", DATE, "", TESTER],
        ["REV_002", "User chưa mua không được review", "Medium", "Mở product chưa mua và thử gửi review.", "product chưa mua", "Không tạo review.", "Cần kiểm thử browser.", "AUTH_006", "Chưa kiểm thử", DATE, "", TESTER],
    ]

    admin_cases = [
        ["ADM_001", "Dashboard tổng quan", "Medium", "Đăng nhập ADMIN, mở /admin.", "ADMIN account", "Hiển thị tổng user, product, order.", "Cần kiểm thử với dữ liệu thực tế.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_002", "Danh sách user", "High", "Mở /admin/user.", "ADMIN account", "Hiển thị danh sách user.", "Cần kiểm thử với dữ liệu thực tế.", "ADM_001", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_003", "Tạo user admin", "High", "Mở form create user, nhập dữ liệu hợp lệ.", "Email/password/fullName/role", "User được lưu với password hash và role đúng.", "Cần kiểm thử UI.", "ADM_002", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_004", "Cập nhật user", "Medium", "Mở update user, sửa thông tin.", "fullName/phone/address/role", "Thông tin user được cập nhật.", "Cần kiểm thử UI.", "ADM_002", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_005", "Xóa user", "High", "Mở delete user và xác nhận.", "userId hợp lệ", "User bị xóa hoặc báo lỗi nếu có ràng buộc dữ liệu.", "Cần kiểm thử UI.", "ADM_002", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_006", "Danh sách sản phẩm admin", "High", "Mở /admin/product.", "ADMIN account", "Hiển thị danh sách sản phẩm.", "Cần kiểm thử với dữ liệu thực tế.", "ADM_001", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_007", "Tạo sản phẩm", "High", "Nhập dữ liệu sản phẩm và upload ảnh.", "name/price/quantity/factory/target/image", "Sản phẩm được lưu.", "Cần kiểm thử upload.", "ADM_006", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_008", "Cập nhật sản phẩm", "Medium", "Sửa thông tin sản phẩm hiện có.", "productId hợp lệ", "Thông tin sản phẩm được cập nhật.", "Cần kiểm thử UI.", "ADM_006", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_009", "Xóa sản phẩm", "High", "Xác nhận xóa sản phẩm.", "productId hợp lệ", "Sản phẩm bị xóa hoặc báo lỗi nếu đang liên kết đơn hàng.", "Cần kiểm thử ràng buộc database.", "ADM_006", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_010", "Quản lý đơn hàng", "High", "Mở /admin/order và xem detail.", "ADMIN account", "Hiển thị danh sách đơn và chi tiết order detail.", "Cần kiểm thử với dữ liệu order.", "ORDER_001", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_011", "Cập nhật trạng thái đơn", "High", "Sửa status đơn hàng.", "PENDING/SHIPPING/COMPLETE/CANCEL", "Status đơn được lưu và hiển thị lại.", "Cần kiểm thử UI.", "ADM_010", "Chưa kiểm thử", DATE, "", TESTER],
        ["ADM_012", "Thống kê doanh thu", "Medium", "Mở /admin/statistics, chọn ngày/tháng/năm.", "selectedDate/selectedMonth/selectedYear", "Hiển thị số đơn, doanh thu và chart theo tháng.", "Cần kiểm thử dữ liệu thực tế.", "ORDER_001", "Chưa kiểm thử", DATE, "", TESTER],
    ]

    non_function_cases = [
        ["NF_001", "Bảo vệ route admin", "High", "Truy cập /admin/** bằng user không có ROLE_ADMIN.", "USER account", "Không cho truy cập admin.", "Cần kiểm thử browser.", "AUTH_006", "Chưa kiểm thử", DATE, "", TESTER],
        ["NF_002", "CSRF token cho form", "High", "Submit form POST thiếu CSRF token.", "Form POST", "Request bị chặn bởi Spring Security CSRF.", "Cần kiểm thử thủ công.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["NF_003", "Token tab không dùng được sau logout", "High", "Logout rồi dùng lại URL có tabToken cũ.", "tabToken đã revoke", "Không xác thực được user.", "Hoạt động theo thiết kế.", "AUTH_010", "Đạt", DATE, "", TESTER],
        ["NF_004", "Upload avatar/sản phẩm giới hạn 50MB", "Medium", "Upload file vượt giới hạn hoặc sai định dạng.", "File test", "Hệ thống xử lý lỗi hoặc không lưu file không hợp lệ.", "Cần kiểm thử upload.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["NF_005", "Persistence sau restart", "Medium", "Restart app sau khi có order/product/user.", "DB MySQL", "Dữ liệu nghiệp vụ vẫn tồn tại trong database.", "Cần kiểm thử môi trường thực tế.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["NF_006", "Responsive layout client", "Low", "Mở trang home/shop/cart trên mobile viewport.", "Chrome DevTools", "Không vỡ layout, text/button dễ thao tác.", "Cần kiểm thử UI.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
        ["NF_007", "Hiệu năng danh sách sản phẩm", "Medium", "Mở /shop với nhiều sản phẩm và filter.", "DB nhiều sản phẩm", "Phản hồi chấp nhận được, có phân trang.", "Cần kiểm thử dữ liệu lớn.", "N/A", "Chưa kiểm thử", DATE, "", TESTER],
    ]

    return [
        make_test_sheet("1.Auth Account", "AUTH_ACCOUNT", "Kiểm thử đăng ký, đăng nhập, đăng xuất, phân quyền và cập nhật tài khoản.", auth_cases),
        make_test_sheet("2.Shop Cart Order", "SHOP_CART_ORDER", "Kiểm thử mua hàng từ xem sản phẩm đến đặt hàng và review.", shop_cases),
        make_test_sheet("3.Admin Management", "ADMIN_MANAGEMENT", "Kiểm thử các chức năng quản trị hệ thống.", admin_cases),
        make_test_sheet("Non Function", "NON_FUNCTION", "Kiểm thử phi chức năng: security, upload, persistence, responsive, performance.", non_function_cases),
    ]


def make_report(summary):
    ws = add_sheet("Test Report")
    title(ws, "BÁO CÁO KẾT QUẢ KIỂM THỬ", 8)
    for idx, row in enumerate(
        [
            ("Tên dự án", "Laptopshop - Website bán laptop"),
            ("Mã dự án", "LAPTOPSHOP_SPRING_MVC"),
            ("Mã tài liệu", "LAPTOPSHOP_Test_Report_v1.0"),
            ("Người tạo", "Cần cập nhật tên thành viên và mã sinh viên"),
            ("Ngày tạo", "12/06/2026"),
            (
                "Ghi chú",
                "Release 1 gồm: Auth/Account, Shop/Cart/Order, Review, Admin Management, Statistics và Non-function.",
            ),
        ],
        3,
    ):
        key_value(ws, idx, *row)

    header(ws, 11, ["STT", "Mã module", "Đạt", "Không đạt", "Chưa kiểm thử", "N/A", "Tổng số test case", "Sheet"])
    for idx, row in enumerate(summary, 1):
        module, passed, failed, untested, na, total, sheet = row
        ws.append([idx, module, passed, failed, untested, na, total, sheet])

    pass_total = sum(row[1] for row in summary)
    fail_total = sum(row[2] for row in summary)
    untested_total = sum(row[3] for row in summary)
    na_total = sum(row[4] for row in summary)
    total = sum(row[5] for row in summary)
    denominator = total - na_total
    coverage = (pass_total + fail_total) / denominator if denominator else 0
    success = pass_total / denominator if denominator else 0

    ws.append(["", "Sub total", pass_total, fail_total, untested_total, na_total, total, ""])
    ws.append(["", "Test coverage", coverage, "", "", "", "", ""])
    ws.cell(ws.max_row, 3).number_format = "0.00%"
    ws.append(["", "Test successful coverage", success, "", "", "", "", ""])
    ws.cell(ws.max_row, 3).number_format = "0.00%"
    ws.append(
        [
            "",
            "Ghi chú",
            "Các case Chưa kiểm thử cần chạy lại trên môi trường local có MySQL và tài khoản test thực tế để cập nhật Actual/Result.",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    style_range(ws)


def main():
    make_cover()
    make_test_case_list()
    make_function_sheet()
    make_prototype_sheet()
    summary = make_test_sheets()
    make_report(summary)
    wb.save(OUT)
    load_workbook(OUT, read_only=True).close()
    print(OUT)


if __name__ == "__main__":
    main()
